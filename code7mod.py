import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import requests
from bs4 import BeautifulSoup

'''
def export_to_excel():
    patent_numbers = text_input.get("1.0", tk.END).strip().split("\n")
    if not patent_numbers or not any(patent_numbers):
        messagebox.showwarning("No Input", "Please enter at least one patent or publication number.")
        return

    wb = Workbook()
    ws = wb.active

    # Set headers
    headers = ["Ref No.", "Google Link", "Espacenet Link", "USPTO Link", "ABSTRACT"]
    column_widths = [25, 40, 40, 40, 70]
    ws.append(headers)
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+col_num)].width = width

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for number in patent_numbers:
        cleaned_number = number.strip()
        if not cleaned_number:
            continue
        
        # Prepare hyperlinks
        google_url = f"https://patents.google.com/patent/{cleaned_number}"
        espacenet_url = f"https://worldwide.espacenet.com/patent/search?q={cleaned_number}"
        uspto_number = cleaned_number.replace("US", "")
        uspto_url = f"https://ppubs.uspto.gov/pubwebapp/external.html?q={uspto_number}.pn."
        
        ws.append([            cleaned_number,
            f'=HYPERLINK("{google_url}", "{cleaned_number}")',
            f'=HYPERLINK("{espacenet_url}", "{cleaned_number}")',
            f'=HYPERLINK("{uspto_url}", "{cleaned_number}")'
        ])

    # Timestamped filename
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"patent_export_{now}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Data exported to {filename}")
'''
def get_abstract(google_url):
    try:
        response = requests.get(google_url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            # Abstract is in <meta name="description" content="...">
            meta = soup.find('meta', attrs={'name': 'description'})
            if meta:
                return meta.get('content', '').strip()
        return "Abstract not found"
    except Exception as e:
        return f"Error: {e}"

def export_to_excel():
    patent_numbers = text_input.get("1.0", tk.END).strip().split("\n")
    if not patent_numbers or not any(patent_numbers):
        messagebox.showwarning("No Input", "Please enter at least one patent or publication number.")
        return

    wb = Workbook()
    ws = wb.active

    headers = ["Ref No.", "Google Link", "Espacenet Link", "USPTO Link", "ABSTRACT"]
    column_widths = [25, 40, 40, 40, 70]
    ws.append(headers)
    
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+col_num)].width = width

 # Set column widths and wrap text for abstract column
    for col_num, width in enumerate(column_widths, 1):
        column_letter = chr(64+col_num)
        ws.column_dimensions[column_letter].width = width
        if col_num == 5:  # Abstract column (E)
            for cell in ws[column_letter]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # Optional: Set row height to accommodate wrapped text
                ws.row_dimensions[ws.max_row].height = 45  # Adjust this value as needed

#STYLE HEADER
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for number in patent_numbers:
        cleaned_number = number.strip()
        if not cleaned_number:
            continue

        google_url = f"https://patents.google.com/patent/{cleaned_number}"
        espacenet_url = f"https://worldwide.espacenet.com/patent/search?q={cleaned_number}"
        uspto_number = cleaned_number.replace("US", "")
        uspto_url = f"https://ppubs.uspto.gov/pubwebapp/external.html?q={uspto_number}.pn."

        abstract = get_abstract(google_url)

        ws.append([            cleaned_number,
            f'=HYPERLINK("{google_url}", "{cleaned_number}")',
            f'=HYPERLINK("{espacenet_url}", "{cleaned_number}")',
            f'=HYPERLINK("{uspto_url}", "{cleaned_number}")',
            abstract
        ])

    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"patent_export_{now}.xlsx"
    wb.save(filename)
    messagebox.showinfo("Export Successful", f"Data exported to {filename}")

# Tkinter UI
root = tk.Tk()
root.title("Patent Exporter")

tk.Label(root, text="Enter patent/publication numbers (one per line):").pack(pady=5)
text_input = tk.Text(root, height=15, width=45)
text_input.pack(padx=10)

tk.Button(root, text="Export to Excel", command=export_to_excel).pack(pady=10)

root.mainloop()