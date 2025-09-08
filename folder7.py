import os
import sqlite3
from pathlib import Path
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
 


# Third-party modules
try:
    from tkcalendar import DateEntry
except ImportError:
    raise SystemExit("tkcalendar is required. Install it with: pip install tkcalendar")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font  
except ImportError:
    raise SystemExit("openpyxl is required. Install it with: pip install openpyxl")


# -----------------------
# Paths and persistence
# -----------------------
def get_base_dir():
    documents = Path.home() / "desktop"
    base = desktop / "clients"
    base.mkdir(parents=True, exist_ok=True)
    return base

BASE_DIR = get_base_dir()
DB_PATH = BASE_DIR / "clients.db"
EXCEL_PATH = BASE_DIR / "clients.xlsx"


# -----------------------
# Utilities
# -----------------------
INVALID_CHARS = set(r'\/:*?"<>|')

def sanitize_for_folder(name: str) -> str:
    # Strip, replace invalid characters, collapse spaces
    name = name.strip()
    cleaned = []
    for ch in name:
        if ch in INVALID_CHARS:
            cleaned.append("_")
        else:
            cleaned.append(ch)
    safe = "".join(cleaned)
    # Replace whitespace runs with single underscore
    safe = "_".join(safe.split())
    return safe[:150] if len(safe) > 150 else safe  # keep reasonable length

def unique_folder_path(base: Path, folder_name: str) -> Path:
    candidate = base / folder_name
    counter = 1
    while candidate.exists():
        candidate = base / f"{folder_name}_{counter}"
        counter += 1
    return candidate

def path_as_uri(p: Path) -> str:
    # Convert local path to file:// URI for Excel hyperlinks
    return p.resolve().as_uri()

def open_in_explorer(path: Path):
    try:
        if os.name == "nt":
            os.startfile(str(path))  # Windows
        elif sys.platform == "darwin":
            import subprocess
            subprocess.run(["open", str(path)], check=False)
        else:
            import subprocess
            subprocess.run(["xdg-open", str(path)], check=False)
    except Exception:
        pass


# -----------------------
# Database
# -----------------------
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS matters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                matter_no TEXT NOT NULL,
                date_due TEXT NOT NULL,
                folder_path TEXT NOT NULL,
                netdocs_path TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.commit()

def insert_record(filename: str, matter_no: str, date_due: str, folder_path: str, netdocs_path: str,):
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO matters (filename, matter_no, date_due, folder_path, netdocs_path, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (filename, matter_no, date_due, folder_path, netdocs_path, datetime.utcnow().isoformat(timespec="seconds"))
        )
        conn.commit()

def fetch_all_records():
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT filename, matter_no, date_due, folder_path, netdocs_path
            FROM matters
            ORDER BY datetime(created_at) ASC
            """
        )
        return cur.fetchall()


# -----------------------
# Excel export
# -----------------------
def export_to_excel():
    rows = fetch_all_records()
    if not rows:
        messagebox.showinfo("Export", "No records to export yet.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"


  # Set headers
    headers = ["Filename", "Matter No.", "Date_Due", "Link to folder", "Netdocs_Path"]
    column_widths = [40, 20, 20, 60, 70]
    ws.append(headers)
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+col_num)].width = width

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r_idx, row in enumerate(rows, start=2):
        filename, matter_no, date_due, folder_path, netdocs_path = row
        ws.cell(row=r_idx, column=1, value=filename)
        ws.cell(row=r_idx, column=2, value=matter_no)
        ws.cell(row=r_idx, column=3, value=date_due)


        link_cell = ws.cell(row=r_idx, column=4, value=str(folder_path))
        try:
            link_cell.hyperlink = path_as_uri(Path(folder_path))
            link_cell.style = "Hyperlink"
        except Exception:
            # Fallback: leave as plain text if hyperlink cannot be formed
            pass
        link_cell2=ws.cell(row=r_idx, column=5, value=str(netdocs_path))
        try:
            link_cell2.hyperlink=(netdocs_path)
            link_cell2.style= "Hyperlink"
        except Exception:
            # Fallback: leave as plain text if hyperlink cannot be formed
            pass
    try:
        wb.save(EXCEL_PATH)
        messagebox.showinfo("Export", f"Exported to:\n{EXCEL_PATH}")
    except PermissionError:
        messagebox.showerror("Export error", "Excel file is open. Please close it and try again.")
    except Exception as e:
        messagebox.showerror("Export error", f"Could not save Excel:\n{e}")


# -----------------------
# Tkinter UI
# -----------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Clients – Folder + DB + Excel")
        self.geometry("520x260")
        self.resizable(False, False)

        self.create_widgets()

    def create_widgets(self):
        pad = {"padx": 10, "pady": 6}

        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, **pad)

        # Filename
        ttk.Label(frm, text="Filename:").grid(row=0, column=0, sticky="e", **pad)
        self.filename_var = tk.StringVar()
        self.filename_entry = ttk.Entry(frm, textvariable=self.filename_var, width=40)
        self.filename_entry.grid(row=0, column=1, columnspan=2, sticky="we", **pad)

        # Matter No.
        ttk.Label(frm, text="Matter No.:").grid(row=1, column=0, sticky="e", **pad)
        self.matter_var = tk.StringVar()
        self.matter_entry = ttk.Entry(frm, textvariable=self.matter_var, width=40)
        self.matter_entry.grid(row=1, column=1, columnspan=2, sticky="we", **pad)

        # Date Due (calendar)
        ttk.Label(frm, text="Date_Due:").grid(row=2, column=0, sticky="e", **pad)
        self.date_entry = DateEntry(frm, date_pattern="yyyy-mm-dd", width=18)
        self.date_entry.grid(row=2, column=1, sticky="w", **pad)

        # Netdocs_path
        ttk.Label(frm, text="Netdocs_Path:").grid(row=3, column=0, sticky="e", **pad)
        self.netdocs_var = tk.StringVar()
        self.netdocs_entry = ttk.Entry(frm, textvariable=self.netdocs_var, width=60)
        self.netdocs_entry.grid(row=3, column=1, columnspan=2, sticky="we", **pad)

        # Buttons
        self.save_btn = ttk.Button(frm, text="Save + Create Folder", command=self.save_record)
        self.save_btn.grid(row=4, column=1, sticky="w", **pad)

        self.export_btn = ttk.Button(frm, text="Export to Excel", command=export_to_excel)
        self.export_btn.grid(row=4, column=2, sticky="w", **pad)

        self.open_base_btn = ttk.Button(frm, text="Open Clients Folder", command=lambda: open_in_explorer(BASE_DIR))
        self.open_base_btn.grid(row=5, column=1, sticky="w", **pad)

        self.quit_btn = ttk.Button(frm, text="Quit", command=self.destroy)
        self.quit_btn.grid(row=6, column=2, sticky="w", **pad)

        # Status
        self.status = tk.StringVar(value=f"Base directory: {BASE_DIR}")
        ttk.Label(frm, textvariable=self.status, foreground="#555").grid(row=5, column=0, columnspan=3, sticky="w", **pad)

        # Column weights 
        frm.grid_columnconfigure(1, weight=1)

    def save_record(self):
        filename = self.filename_var.get().strip()
        matter_no = self.matter_var.get().strip()
        date_due = self.date_entry.get_date().strftime("%Y-%m-%d")
        netdocs_path = self.netdocs_var.get().strip()

        if not filename or not matter_no:
            messagebox.showwarning("Validation", "Please fill in Filename and Matter No.")
            return

        safe_filename = sanitize_for_folder(filename)
        safe_matter = sanitize_for_folder(matter_no)
        safe_date = sanitize_for_folder(date_due)

        folder_name = f"{safe_filename}_{safe_matter}_{safe_date}"
        folder_path = unique_folder_path(BASE_DIR, folder_name)
        try:
            folder_path.mkdir(parents=True, exist_ok=False)
        except Exception as e:
            messagebox.showerror("Folder error", f"Could not create folder:\n{e}")
            return

        try:
            insert_record(filename, matter_no, date_due,  str(folder_path), netdocs_path)
        except Exception as e:
            messagebox.showerror("Database error", f"Could not write to database:\n{e}")
            return

        self.status.set(f"Saved and created folder: {folder_path}")
        messagebox.showinfo("Success", "Record saved and folder created ✅")

        # Optional: clear fields
        # self.filename_var.set("")
        # self.matter_var.set("")


if __name__ == "__main__":
    init_db()
    app = App()
    app.mainloop()