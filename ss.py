import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
import os
import re
from urllib.parse import urljoin
import time

class PatentScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Google Patents Scraper")
        self.root.geometry("600x500")
        
        # Variables
        self.excel_file = tk.StringVar(value="patents_data.xlsx")
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title_label = ttk.Label(main_frame, text="Google Patents Scraper", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Patent number input
        ttk.Label(main_frame, text="Patent Number:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.patent_entry = ttk.Entry(main_frame, width=30)
        self.patent_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        
        # Example label
        example_label = ttk.Label(main_frame, text="Example: US10123456B2", 
                                 font=("Arial", 8), foreground="gray")
        example_label.grid(row=2, column=1, sticky=tk.W, padx=(10, 0))
        
        # Excel file selection
        ttk.Label(main_frame, text="Excel File:").grid(row=3, column=0, sticky=tk.W, pady=(20, 5))
        
        file_frame = ttk.Frame(main_frame)
        file_frame.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=(20, 5), padx=(10, 0))
        
        self.file_entry = ttk.Entry(file_frame, textvariable=self.excel_file, width=25)
        self.file_entry.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        browse_btn = ttk.Button(file_frame, text="Browse", command=self.browse_file)
        browse_btn.grid(row=0, column=1, padx=(5, 0))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=(30, 0))
        
        self.scrape_btn = ttk.Button(button_frame, text="Scrape Patent", 
                                    command=self.scrape_patent, style="Accent.TButton")
        self.scrape_btn.grid(row=0, column=0, padx=(0, 10))
        
        view_btn = ttk.Button(button_frame, text="View Excel File", 
                             command=self.view_excel)
        view_btn.grid(row=0, column=1)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(20, 0))
        
        # Status text
        self.status_text = tk.Text(main_frame, height=12, width=60, wrap=tk.WORD)
        self.status_text.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(20, 0))
        
        # Scrollbar for status text
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.status_text.yview)
        scrollbar.grid(row=6, column=2, sticky=(tk.N, tk.S), pady=(20, 0))
        self.status_text.configure(yscrollcommand=scrollbar.set)
        
        # Configure grid weights
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(6, weight=1)
        file_frame.columnconfigure(0, weight=1)
        
    def browse_file(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.excel_file.set(filename)
    
    def log_message(self, message):
        self.status_text.insert(tk.END, f"{message}\n")
        self.status_text.see(tk.END)
        self.root.update()
    
    def clean_patent_number(self, patent_num):
        # Remove any spaces and convert to uppercase
        return re.sub(r'[^\w]', '', patent_num.upper())
    
    def scrape_patent(self):
        patent_num = self.patent_entry.get().strip()
        if not patent_num:
            messagebox.showerror("Error", "Please enter a patent number")
            return
        
        # Disable button and start progress
        self.scrape_btn.configure(state='disabled')
        self.progress.start()
        
        try:
            self.log_message(f"Starting scrape for patent: {patent_num}")
            
            # Clean patent number
            clean_patent = self.clean_patent_number(patent_num)
            
            # Construct Google Patents URL
            url = f"https://patents.google.com/patent/{clean_patent}"
            self.log_message(f"URL: {url}")
            
            # Set up headers to mimic a real browser
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            # Make request
            self.log_message("Fetching patent page...")
            response = requests.get(url, headers=headers, timeout=30)
            
            if response.status_code != 200:
                raise Exception(f"Failed to fetch page. Status code: {response.status_code}")
            
            # Parse HTML
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extract abstract
            abstract = self.extract_abstract(soup)
            
            if not abstract:
                self.log_message("Warning: Could not find abstract")
                abstract = "Abstract not found"
            
            # Save to Excel
            self.save_to_excel(patent_num, url, abstract)
            
            self.log_message("Successfully scraped patent data!")
            messagebox.showinfo("Success", f"Patent {patent_num} data saved to Excel!")
            
        except Exception as e:
            error_msg = f"Error scraping patent: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("Error", error_msg)
        
        finally:
            # Re-enable button and stop progress
            self.scrape_btn.configure(state='normal')
            self.progress.stop()
    
    def extract_abstract(self, soup):
        # Try multiple selectors for abstract
        abstract_selectors = [
            'section[data-section-id="abstract"] div.abstract',
            'section[data-section="abstract"] div.abstract',
            '.abstract',
            '[data-section-id="abstract"]',
            'section:contains("Abstract")',
        ]
        
        for selector in abstract_selectors:
            try:
                if ':contains' in selector:
                    # Handle special case for text search
                    sections = soup.find_all('section')
                    for section in sections:
                        if 'abstract' in section.get_text().lower():
                            abstract_div = section.find('div')
                            if abstract_div:
                                return abstract_div.get_text().strip()
                else:
                    element = soup.select_one(selector)
                    if element:
                        text = element.get_text().strip()
                        if text and len(text) > 20:  # Ensure it's substantial
                            return text
            except Exception as e:
                self.log_message(f"Selector failed: {selector} - {e}")
                continue
        
        # Fallback: look for any div containing "abstract" in class or text
        abstract_divs = soup.find_all('div', class_=lambda x: x and 'abstract' in x.lower())
        for div in abstract_divs:
            text = div.get_text().strip()
            if text and len(text) > 20:
                return text
        
        # Last resort: search for text patterns
        text = soup.get_text()
        abstract_match = re.search(r'Abstract\s*[:.]?\s*([^.]{50,500})', text, re.IGNORECASE | re.DOTALL)
        if abstract_match:
            return abstract_match.group(1).strip()
        
        return None
    
    def save_to_excel(self, patent_num, url, abstract):
        filename = self.excel_file.get()
        
        # Check if file exists
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Add headers
            ws['A1'] = 'Patent Number'
            ws['B1'] = 'Google Patents URL'
            ws['C1'] = 'Abstract'
            ws['D1'] = 'Date Scraped'
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Add data
        ws[f'A{next_row}'] = patent_num
        ws[f'B{next_row}'] = url
        ws[f'C{next_row}'] = abstract
        ws[f'D{next_row}'] = time.strftime('%Y-%m-%d %H:%M:%S')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 20
        
        # Wrap text in abstract column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        # Save file
        wb.save(filename)
        self.log_message(f"Data saved to: {filename}")
    
    def view_excel(self):
        filename = self.excel_file.get()
        if os.path.exists(filename):
            try:
                # Try to open with default program
                if os.name == 'nt':  # Windows
                    os.startfile(filename)
                elif os.name == 'posix':  # macOS and Linux
                    os.system(f'open "{filename}"')
            except Exception as e:
                messagebox.showerror("Error", f"Could not open file: {e}")
        else:
            messagebox.showwarning("Warning", "Excel file does not exist yet")

def main():
    root = tk.Tk()
    app = PatentScraperApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()