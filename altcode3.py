import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import requests
from bs4 import BeautifulSoup

def get_patent_details(google_url):
    try:
        response = requests.get(google_url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Extract Title
            title_tag = soup.find('meta', attrs={'name': 'DC.title'})
            title = title_tag.get('content', '').strip() if title_tag else "Title not found"

            # Extract Inventor
            inventor_tag = soup.find('meta', attrs={'name': 'DC.contributor'})
            inventor = inventor_tag.get('content', '').strip() if inventor_tag else "Inventor not found"

            # Extract Publication Date
            pub_date_tag = soup.find('meta', attrs={'name': 'DC.date'})
            publication_date = pub_date_tag.get('content', '').strip() if pub_date_tag else "Publication date not found"

            # Extract Abstract
            meta = soup.find('meta', attrs={'name': 'DC.description'})
            abstract = meta.get('content', '').strip() if meta else "Abstract not found"

            return title, inventor, publication_date, abstract
        return None
    except Exception as e:
        return None

def export_to_excel():
    patent_numbers = text_input.get("1.0", tk.END).strip().split("\n")
    if not patent_numbers or not any(patent_numbers):
        messagebox.showwarning("No Input", "Please enter at least one patent or publication number.")
        return

    wb = Workbook()
    ws = wb.active

    # Set headers
    headers = ["Ref No.", "Google Link", "Espacenet Link", "USPTO Link", "Title", "Inventor", "Publication Date", "Abstract"]
    column_widths = [25, 40, 40, 40, 30, 30, 20, 70]
    ws.append(headers)

    # Set column widths and wrap text for abstract column
    for col_num, width in enumerate(column_widths, 1):
        column_letter = chr(64+col_num)
        ws.column_dimensions[column_letter].width = width
        if col_num == 8:  # Abstract column (H)
            for cell in ws[column_letter]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for number in patent_numbers:
        cleaned_number = number.strip()
        if not cleaned_number:
            continue

        google_url = f"https://patents.google.com/patent/{cleaned_number}"
        espacenet_url = f"https://worldwide.espacenet.com/patent/search?q={cleaned_number}"
        uspto_number = cleaned_number.replace("US", "")
        uspto_url = f"https://ppubs.uspto.gov/pubwebapp/external.html?q={uspto_number}.pn."

        # Get patent details
        details = get_patent_details(google_url)
        if details:
            title, inventor, publication_date, abstract = details

            # Add row
            ws.append([
                cleaned_number,
                f'=HYPERLINK("{google_url}", "{cleaned_number}")',
                f'=HYPERLINK("{espacenet_url}", "{cleaned_number}")',
                f'=HYPERLINK("{uspto_url}", "{cleaned_number}")',
                title,
                inventor,
                publication_date,
                abstract
            ])

    # Save the workbook
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"patent_export_{now}.xlsx"
    wb.save(filename)

    messagebox.showinfo("Export Successful", f"Data exported to {filename}")

# Tkinter UI
root = tk.Tk()
root.title("Patent Exporter")

tk.Label(root, text="Enter patent/publication numbers (one per line):").pack(pady=5)
text_input = tk.Text(root, height=15, width=45)
text_input.pack(padx=10)

tk.Button(root, text="Export to Excel", command=export_to_excel).pack(pady=10)

root.mainloop()